"""Export des écritures validées vers un .xlsx importable dans Sage.

Colonnes Sage : journal, date, compte_debit, compte_credit, libelle, debit, credit, piece
"""

from __future__ import annotations

import io
import logging
import os
from typing import Any, TypedDict

import openpyxl
from openpyxl.styles import Font

__all__ = [
    "SageExporterResult",
    "build_export_rows",
    "write_xlsx",
    "run",
]

logger = logging.getLogger(__name__)

SAGE_COLUMNS = ["journal", "date", "compte_debit", "compte_credit", "libelle", "debit", "credit", "piece"]
STORAGE_BUCKET = "exports"


class SageExporterResult(TypedDict):
    dossier_id: str
    mois: str
    exported: bool
    nb_ecritures: int
    storage_path: str | None
    journal_id: str | None
    erreur: str | None


def _get_supabase():  # type: ignore[return]
    from supabase import create_client  # noqa: PLC0415
    url = os.getenv("SUPABASE_URL", "")
    key = os.getenv("SUPABASE_SERVICE_KEY", "")
    if not url or not key:
        raise ValueError("SUPABASE_URL et SUPABASE_SERVICE_KEY requis")
    return create_client(url, key)


def build_export_rows(supabase, dossier_id: str, periode: str) -> list[dict[str, Any]]:
    """Retourne uniquement les écritures statut='valide' pour dossier+période.

    periode: 'YYYY-MM' — filtre sur date (colonnes date LIKE 'YYYY-MM%').
    """
    try:
        resp = (
            supabase.table("ecritures")
            .select(
                "id, journal, date_ecriture, compte_debit, compte_credit, libelle, "
                "montant, reference, montant_ttc, statut"
            )
            .eq("dossier_id", dossier_id)
            .eq("statut", "valide")
            .like("date_ecriture", f"{periode}%")
            .order("date_ecriture")
            .execute()
        )
    except Exception as exc:
        raise RuntimeError(f"build_export_rows — lecture échouée : {exc}") from exc

    rows = []
    for e in resp.data or []:
        rows.append({
            "journal": e.get("journal") or "",
            "date": e.get("date_ecriture") or "",
            "compte_debit": e.get("compte_debit") or "",
            "compte_credit": e.get("compte_credit") or "",
            "libelle": e.get("libelle") or "",
            "debit": e.get("montant"),
            "credit": e.get("montant"),
            "piece": e.get("reference") or "",
        })
    return rows


def write_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Sérialise les lignes en .xlsx Sage et retourne les bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Sage"

    # Header
    for col_idx, col_name in enumerate(SAGE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(SAGE_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run(
    dossier_id: str,
    mois: str,
    dry_run: bool = False,
    supabase=None,
) -> SageExporterResult:
    """Exporte les écritures validées d'un dossier/mois en .xlsx Sage.

    Args:
        dossier_id: UUID du dossier.
        mois: Période 'YYYY-MM'.
        dry_run: Si True, calcule sans uploader ni logguer.
        supabase: Client optionnel (construit depuis env sinon).

    Returns:
        SageExporterResult.
    """
    logger.info(f"sage_exporter — dossier {dossier_id}, mois {mois}")

    try:
        sb = supabase or _get_supabase()
    except Exception as exc:
        return SageExporterResult(
            dossier_id=dossier_id, mois=mois, exported=False, nb_ecritures=0,
            storage_path=None, journal_id=None, erreur=str(exc),
        )

    try:
        rows = build_export_rows(sb, dossier_id, mois)
    except Exception as exc:
        return SageExporterResult(
            dossier_id=dossier_id, mois=mois, exported=False, nb_ecritures=0,
            storage_path=None, journal_id=None, erreur=str(exc),
        )

    if not rows:
        logger.info(f"sage_exporter — aucune écriture valide pour {dossier_id}/{mois}")
        return SageExporterResult(
            dossier_id=dossier_id, mois=mois, exported=False, nb_ecritures=0,
            storage_path=None, journal_id=None, erreur=None,
        )

    try:
        xlsx_bytes = write_xlsx(rows)
    except Exception as exc:
        return SageExporterResult(
            dossier_id=dossier_id, mois=mois, exported=False, nb_ecritures=len(rows),
            storage_path=None, journal_id=None, erreur=f"write_xlsx : {exc}",
        )

    storage_path = f"{dossier_id}/{mois}_sage_export.xlsx"
    journal_id: str | None = None

    if not dry_run:
        try:
            sb.storage.from_(STORAGE_BUCKET).upload(
                path=storage_path,
                file=xlsx_bytes,
                file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            )
        except Exception as exc:
            return SageExporterResult(
                dossier_id=dossier_id, mois=mois, exported=False, nb_ecritures=len(rows),
                storage_path=None, journal_id=None, erreur=f"upload Storage : {exc}",
            )

        try:
            resp = sb.table("journaux").insert({
                "dossier_id": dossier_id,
                "type_action": "sage_saisie",
                "statut": "ok",
                "contenu": f"Export Sage {mois} — {len(rows)} écriture(s)",
                "metadata": {"mois": mois, "nb_ecritures": len(rows), "storage_path": storage_path},
            }).execute()
            if resp.data:
                journal_id = resp.data[0]["id"]
        except Exception as exc:
            logger.warning(f"sage_exporter — log journal échoué : {exc}")

    return SageExporterResult(
        dossier_id=dossier_id, mois=mois, exported=True, nb_ecritures=len(rows),
        storage_path=storage_path, journal_id=journal_id, erreur=None,
    )
