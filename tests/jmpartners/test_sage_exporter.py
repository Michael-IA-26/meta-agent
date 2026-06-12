"""Tests TDD — sage_exporter (zero réseau)."""

from __future__ import annotations

import io
from unittest.mock import MagicMock

import openpyxl

from apps.jmpartners.agents.sage_exporter import (
    SAGE_COLUMNS,
    build_export_rows,
    run,
    write_xlsx,
)


def _sb_with_ecritures(rows):
    sb = MagicMock()
    (
        sb.table.return_value
        .select.return_value
        .eq.return_value
        .eq.return_value
        .like.return_value
        .order.return_value
        .execute.return_value
        .data
    ) = rows
    sb.storage.from_.return_value.upload.return_value = {}
    sb.table.return_value.insert.return_value.execute.return_value.data = [{"id": "j-1"}]
    return sb


def _ecriture(
    statut="valide",
    date="2026-05-15",
    journal="ACH",
    reference="FA-001",
    debit=1200.0,
    credit=None,
    compte_debit="6070",
    compte_credit="401",
    libelle="Achat",
):
    return {
        "id": "ecr-1",
        "journal": journal,
        "date": date,
        "compte_debit": compte_debit,
        "compte_credit": compte_credit,
        "libelle": libelle,
        "debit": debit,
        "credit": credit,
        "reference": reference,
        "montant_ttc": debit,
        "statut": statut,
    }


# ── build_export_rows ─────────────────────────────────────────────────────────

def test_build_export_rows_only_valide():
    """build_export_rows returns only statut='valide' entries."""
    sb = _sb_with_ecritures([_ecriture()])
    rows = build_export_rows(sb, "doss-1", "2026-05")
    assert len(rows) == 1
    assert rows[0]["journal"] == "ACH"
    assert rows[0]["piece"] == "FA-001"


def test_build_export_rows_empty_set():
    sb = _sb_with_ecritures([])
    rows = build_export_rows(sb, "doss-1", "2026-05")
    assert rows == []


def test_build_export_rows_maps_sage_columns():
    """Each row has exactly the SAGE_COLUMNS keys."""
    sb = _sb_with_ecritures([_ecriture()])
    rows = build_export_rows(sb, "doss-1", "2026-05")
    assert set(rows[0].keys()) == set(SAGE_COLUMNS)


def test_build_export_rows_amounts_from_debit_credit():
    """debit/credit columns come from ecriture debit/credit fields."""
    sb = _sb_with_ecritures([_ecriture(debit=1200.0, credit=None)])
    rows = build_export_rows(sb, "doss-1", "2026-05")
    assert rows[0]["debit"] == 1200.0
    assert rows[0]["credit"] is None


# ── write_xlsx ────────────────────────────────────────────────────────────────

def test_write_xlsx_produces_valid_xlsx():
    """write_xlsx returns valid xlsx bytes."""
    rows = [build_export_rows(_sb_with_ecritures([_ecriture()]), "doss-1", "2026-05")[0]]
    xlsx_bytes = write_xlsx(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "JOURNAL"


def test_write_xlsx_data_row():
    """Data row is written after header."""
    rows = [{"journal": "ACH", "date": "2026-05-15", "compte_debit": "6070",
              "compte_credit": "401", "libelle": "Achat", "debit": 1200.0,
              "credit": None, "piece": "FA-001"}]
    xlsx_bytes = write_xlsx(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "ACH"
    assert ws.cell(row=2, column=6).value == 1200.0


# ── run ───────────────────────────────────────────────────────────────────────

def test_run_empty_returns_exported_false():
    """No valid entries → exported=False, no upload."""
    sb = _sb_with_ecritures([])
    result = run("doss-1", "2026-05", supabase=sb)
    assert result["exported"] is False
    assert result["nb_ecritures"] == 0
    sb.storage.from_.assert_not_called()


def test_run_with_data_uploads_to_storage():
    """Entries present → xlsx uploaded to Storage bucket 'exports'."""
    sb = _sb_with_ecritures([_ecriture()])
    result = run("doss-1", "2026-05", supabase=sb)
    assert result["exported"] is True
    assert result["nb_ecritures"] == 1
    sb.storage.from_.assert_called_with("exports")
    sb.storage.from_.return_value.upload.assert_called_once()


def test_run_logs_journaux_row():
    """run() inserts a 'sage_saisie' journaux row."""
    sb = _sb_with_ecritures([_ecriture()])
    result = run("doss-1", "2026-05", supabase=sb)
    assert result["journal_id"] == "j-1"
    insert_call = sb.table.return_value.insert.call_args[0][0]
    assert insert_call["type_action"] == "sage_saisie"


def test_run_dry_run_no_upload_no_log():
    """dry_run=True → xlsx built but not uploaded, no journal row."""
    sb = _sb_with_ecritures([_ecriture()])
    result = run("doss-1", "2026-05", dry_run=True, supabase=sb)
    assert result["exported"] is True
    sb.storage.from_.assert_not_called()
    sb.table.return_value.insert.assert_not_called()


def test_run_storage_path_contains_dossier_and_mois():
    sb = _sb_with_ecritures([_ecriture()])
    result = run("doss-42", "2026-05", supabase=sb)
    assert "doss-42" in (result["storage_path"] or "")
    assert "2026-05" in (result["storage_path"] or "")
